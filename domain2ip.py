import os,sys
import socket
import re
import openpyxl
from openpyxl import Workbook

def showBanner():
    banner = """
    usage: python3 domain2ip.py <filePath>or<dirPath>
    filePath includes xlsx or txt or NULLEXT
                            authored by rhaps
    """
    print(banner)


def dnsProcess(target):
    ipPattern = re.compile(r"\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b")
    res = re.findall(ipPattern, target)
    if not res:
        # 是域名信息，需要解析
        domain = target.lstrip('https//')
        domain = domain.lstrip('http//')
        domain = domain.rstrip('\n')
        if not domain.startswith('www.'):
            domain = 'www.' + domain
        #最后不能有/
        domain = domain.rstrip('/')
        ipList = []
        try:
            print('Processing:%s' % (domain))
            ip = socket.gethostbyname(domain)
            addrs = socket.getaddrinfo(domain, None)
            for item in addrs:
                if item[4][0] not in ipList:
                    ipList.append(item[4][0])
        except Exception as e:
            print(str(e))

    else:
        # 传入的就是IP地址
        ip = re.findall(ipPattern, target)
        ip = ip[0]
        print(ip)
        ipList.append(ip)
    print(ipList)
    return ipList


def xlsxFileProcess(filePathName, filePath, fileName):
    wb = openpyxl.load_workbook(filePathName)	#返回workbook对象
    ws = wb.active
    rowList = list(ws.rows)
    for count, row in enumerate(rowList):
        excelRowCount = count + 1
        urlOrIp = row[0].value
        ipList = dnsProcess(urlOrIp)
        print(ipList)
        try:
            if len(ipList):
                col = 2
                for ip in range(0, len(ipList)):
                    ws.cell(excelRowCount, col, ipList[ip])
                    col = col + 1
        except Exception as e:
            print(str(e))

    wb.save(filePath + '/' + "DNSResult_" + fileName)


def txtFileProcess(filePathName, filePath, fileName):
    #新建一个xlsx文档
    wb = Workbook()
    ws = wb.active
    file = open(filePathName, 'r+')
    row = 1
    while True:
        urlOrIp = file.readline()
        if urlOrIp:
            ws.cell(row, 1, urlOrIp)
            row = row + 1
        else:
            break


    newXlsxFileName = fileName + '.xlsx'
    wb.save(filePath + '/' + newXlsxFileName)
    #处理xlsx文档
    xlsxFileProcess(filePath + '/' + newXlsxFileName, filePath, newXlsxFileName)



def fileProcess(filePathName, filePath, fileName):
    if fileName.endswith('.xlsx'):
        #文件后缀是xlsx，调用处理excel表格的函数
        xlsxFileProcess(filePathName, filePath, fileName)
    else:
        #没有后缀或者后缀为txt,需要先新建一个xlsx文档
        txtFileProcess(filePathName, filePath, fileName)


def dirProcess(targetsDir):
    for filePath, dirNames, fileNames in os.walk(targetsDir):
    	for fileName in fileNames:
            #fileName是单纯的文件名，filePathName是完整的文件路径
            if not fileName.startswith('.'):
                filePathName = os.path.join(os.getcwd(), filePath, fileName)
                print(filePathName)
                print(filePath)
                print(fileName)
                fileProcess(filePathName, filePath, fileName)

if __name__ == "__main__":
    try:
        targets = sys.argv[1]
        while (os.path.exists(targets) == False):
            print('Invalid targetPath\n')
            targets = input("please input valid targetPath")
        if os.path.isdir(targets):
            # 用户输入的是文件夹
            dirProcess(targets)
        elif os.path.isfile(targets):
            # 用户输入的是文件
            targetsFile = targets
            (path, file) = os.path.split(targetsFile)
            if not path:
                path = os.getcwd()
            fileProcess(targetsFile, path, file)
        else:
            # 用户输入的路径无效
            showBanner()
    except Exception as e:
        print(str(e))
        showBanner()







